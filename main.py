import json
from typing import List, Dict, Any
from pprint import pprint
import requests
from openpyxl import Workbook, load_workbook
from config import token, key

class Wildberries:
    URL = 'https://suppliers-api.wildberries.ru'

    def __init__(self, token, key):
        self.token = token
        self.key = key
        self.headers = {
            'content-type': 'application/json',
            "Authorization": self.token
        }
        self.params = {
            'key': self.key
        }

    def _send_post_request(self, endpoint: str, params: dict = None, data: dict = None, files: dict = None,
                           headers: dict = None):
        """Выполнить POST запрос на адрес 'endpoint' с дополнительными заголовками и параметрами"""
        if files is None:
            files = {}
        if headers is None:
            headers = {}
        if params is None:
            params = {}
        if data is None:
            data = ''
        else:
            data = json.dumps(data)
        headers = self.headers | headers
        params = self.params | params
        response = requests.post(self.URL + endpoint, headers=headers, params=params, data=data, files=files)

        return response.json()

    def _send_get_request(self, endpoint: str, params: dict = None, headers: dict = None) -> dict:
        """Выполнить GET запрос на адрес 'endpoint' с дополнительными заголовками и параметрами"""
        if headers is None:
            headers = {}
        if params is None:
            params = {}
        headers = self.headers | headers
        params = self.params | params
        response = requests.get(self.URL + endpoint, headers=headers, params=params)
        if response:
            return response.json()
        else:
            raise Exception((response.status_code, response.text))

    def get_info(self):
        """Получение информации по номенклатурам, их ценам, скидкам и промокодам. Если не указывать фильтры,
        вернётся весь товар."""
        endpoint = '/public/api/v1/info'
        return self._send_get_request(endpoint=endpoint)['result']

    def get_orders(self, dateFrom, flag=0):
        endpoint = '/api/v1/supplier/orders'
        params = {
            'dateFrom': dateFrom,
            'flag': flag
        }
        return self._send_get_request(endpoint=endpoint, params=params)['result']

    def get_sales(self, dateFrom, flag=0):
        endpoint = '/api/v1/supplier/sales'
        params = {
            'dateFrom': dateFrom,
            'flag': flag
        }
        return self._send_get_request(endpoint=endpoint, params=params)['result']

    def get_cards(self, limit: int = 100, offset: int = 0, searchValue: str = '', sortColumn: str = 'updateAt', ascending: bool = True):
        """Получить список карточек поставщика с фильтром и сортировкой"""
        data = {'sort': {
                        'limit': limit,
                        'offset': offset,
                        'searchValue': searchValue,
                        'sortColumn': sortColumn,
                        'ascending': ascending
                        }
                }
        data_str = json.dumps(data)
        response = self._send_post_request(endpoint='/content/v1/cards/list', data=data)

        return response['data']['cards']

    def get_card_imtId(self, imtId: int):
        """Получение карточки поставщика по imt id"""
        data = {"imtID": imtId}
        response = self._send_post_request(endpoint='card/cardByImtID', data=data)

        return response['result']

    def generate_barcode(self, quantity: int) -> List:
        """Позволяет сгенерировать штрих-код для размера"""
        params = {"quantity": quantity}
        response = self._send_request('card/getBarcodes', params=params)

        return response['result']['barcodes']

    def delete_nomenclature(self, nomenclatureID: int):
        """Удаляет одну номенклатуру из карточки товара."""
        params = {"nomenclatureID": nomenclatureID}
        response = self._send_request('card/deleteNomenclature', params=params)

        return response


def set_sheet_from_json(name_sheet: str, data_jison: dict):
    """Заполняет лист Excel данными из json. Если листа нет создает новый"""
    try:
        book = load_workbook('test.xlsx')
    except PermissionError:
        print('Закройте эксель')
    except FileNotFoundError:
        print('Файл не найден, создан новый')
        book = Workbook()
        book.remove(book['Sheet'])
    try:
        sheet = book[name_sheet]
    except KeyError:
        sheet = book.create_sheet(name_sheet)
    headers = list(data_jison[0].keys())
    row = 1
    column = 1
    for header in headers:
        sheet.cell(row=row, column=column).value = header
        column += 1
    row = 2
    for elm in data_jison:
        column = 1
        for data in elm.values():
            if type(data) == dict or type(data) == list:
                data = json.dumps(data, ensure_ascii=False)
            sheet.cell(row=row, column=column).value = data
            column += 1
        row += 1
    book.save('test.xlsx')
    book.close()


def get_json_from_sheet(name_sheet: str) -> list[dict[Any, Any]]:
    """Создает json на основе данных листа Excel."""
    try:
        book = load_workbook('test.xlsx')
    except PermissionError:
        print('Закройте эксель')
    try:
        sheet = book[name_sheet]
    except KeyError:
        print(f'Листа с названием {name_sheet} нет')
        raise Exception
    row_generator = sheet.values
    try:
        headers = next(row_generator)
    except StopIteration:
        print(f'Лист с названием {name_sheet} пустой')
        raise Exception
    data = []
    for row in row_generator:
        dict_row = dict(zip(headers, row))
        data.append(dict_row)
    book.close()

    return data


def main():
    wildberries = Wildberries(token=token, key=key)
    # nomenclature = wildberries.get_info()
    # set_sheet_from_json(name_sheet='nomenclature', data_jison=nomenclature)
    # nomenclature = get_json_from_sheet(name_sheet='nomenclature')
    # print(nomenclature)
    orders = wildberries.get_cards()
    # set_sheet_from_json(name_sheet='orders', data_jison=orders)
    # pprint(orders)
    print(len(orders))

if __name__ == '__main__':
    main()
